# import os
# import pandas as pd
# from datetime import datetime

# # Set folder path
# folder_path = r"C:\Users\hgada\OneDrive - Hearst\Documents\Media_Star_XML\File Output"
# output_diff_file = os.path.join(folder_path, f"MediaStar_Differences_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")

# # Step 1: Find the two most recent .xlsx files
# xlsx_files = [os.path.join(folder_path, f) for f in os.listdir(folder_path) if f.endswith(".xlsx")]
# xlsx_files.sort(key=os.path.getmtime, reverse=True)

# if len(xlsx_files) < 2:
#     raise Exception("Less than two Excel files found in the folder to compare.")

# new_file, old_file = xlsx_files[0], xlsx_files[1]
# print(f"Comparing files:\nNEW: {os.path.basename(new_file)}\nOLD: {os.path.basename(old_file)}")

# # Step 2: Load both files
# new_excel = pd.read_excel(new_file, sheet_name=None)
# old_excel = pd.read_excel(old_file, sheet_name=None)

# # Step 3: Compare sheets
# diff_results = {}

# for sheet in new_excel.keys():
#     if sheet not in old_excel:
#         print(f"Sheet '{sheet}' not found in old file. Skipping.")
#         continue

#     new_df = new_excel[sheet]
#     old_df = old_excel[sheet]

#     if 'eventId' not in new_df.columns or 'eventId' not in old_df.columns:
#         print(f"Sheet '{sheet}' does not contain 'eventId'. Skipping.")
#         continue

#     # Ensure eventId is treated as string for comparison
#     new_df['eventId'] = new_df['eventId'].astype(str)
#     old_df['eventId'] = old_df['eventId'].astype(str)

#     # Find new and missing eventIds
#     new_event_ids = set(new_df['eventId']) - set(old_df['eventId'])
#     missing_event_ids = set(old_df['eventId']) - set(new_df['eventId'])

#     # Extract rows by eventId
#     new_rows = new_df[new_df['eventId'].isin(new_event_ids)]
#     missing_rows = old_df[old_df['eventId'].isin(missing_event_ids)]

#     if not new_rows.empty or not missing_rows.empty:
#         combined = pd.concat([
#             new_rows.assign(ChangeType="Added"),
#             missing_rows.assign(ChangeType="Removed")
#         ])
#         diff_results[sheet] = combined

# # Step 4: Save differences to new Excel file
# if diff_results:
#     with pd.ExcelWriter(output_diff_file, engine="openpyxl") as writer:
#         for sheet, df in diff_results.items():
#             df.to_excel(writer, sheet_name=sheet[:31], index=False)
#     print(f"\n✅ Difference file created at:\n{output_diff_file}")
# else:
#     print("\n✅ No differences found. No file created.")



# import os
# import pandas as pd
# from datetime import datetime

# # Path to your folder
# folder_path = r"C:\Users\hgada\Downloads\Media_Star_XML\File Output"
# timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
# output_file = os.path.join(folder_path, f"Difference_{timestamp}.xlsx")

# # Step 1: Get 2 most recent Excel files
# xlsx_files = [
#     os.path.join(folder_path, f)
#     for f in os.listdir(folder_path)
#     if f.lower().endswith(".xlsx")
# ]
# xlsx_files.sort(key=os.path.getmtime, reverse=True)

# if len(xlsx_files) < 2:
#     raise Exception("Need at least two Excel files to compare.")

# new_file, old_file = xlsx_files[0], xlsx_files[1]
# print(f"Comparing:\nNEW: {os.path.basename(new_file)}\nOLD: {os.path.basename(old_file)}")

# # Step 2: Load both files
# new_excel = pd.read_excel(new_file, sheet_name=None)
# old_excel = pd.read_excel(old_file, sheet_name=None)

# # Step 3: Compare by eventId per common sheet
# differences = []

# for sheet in new_excel:
#     if sheet not in old_excel:
#         continue

#     new_df = new_excel[sheet].copy()
#     old_df = old_excel[sheet].copy()

#     if 'eventId' not in new_df.columns or 'eventId' not in old_df.columns:
#         continue

#     new_df['eventId'] = new_df['eventId'].astype(str)
#     old_df['eventId'] = old_df['eventId'].astype(str)

#     # Set indexes to eventId
#     new_df.set_index('eventId', inplace=True, drop=False)
#     old_df.set_index('eventId', inplace=True, drop=False)

#     # 1. Added eventIds
#     added_ids = set(new_df.index) - set(old_df.index)
#     if added_ids:
#         added = new_df.loc[list(added_ids)].copy()

#         added['ChangeType'] = 'Added'
#         added['Sheet'] = sheet
#         differences.append(added)

#     # 2. Modified eventIds
#     common_ids = set(new_df.index).intersection(old_df.index)
#     for eid in common_ids:
#         new_row = new_df.loc[eid]
#         old_row = old_df.loc[eid]

#         # Check if any value differs
#         if not new_row.equals(old_row):
#             changed = new_row.copy()
#             changed['ChangeType'] = 'Modified'
#             changed['Sheet'] = sheet
#             differences.append(changed.to_frame().T)

# # Step 4: Combine and write to Excel
# if differences:
#     all_diffs = pd.concat(differences, ignore_index=True)
#     with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
#         all_diffs.to_excel(writer, sheet_name="Differences", index=False)
#     print(f"\n✅ Differences written to: {output_file}")
# else:
#     print("\n✅ No differences found between the two files.")


# import os
# import pandas as pd
# from datetime import datetime

# # === Folder paths ===
# input_folder = r"C:\Users\hgada\Downloads\Media_Star_XML\File Output"
# output_folder = r"C:\Users\hgada\Downloads\Media_Star_XML\Differences"
# os.makedirs(output_folder, exist_ok=True)

# # Generate timestamped output file
# timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
# output_file = os.path.join(output_folder, f"Difference_{timestamp}.xlsx")

# # === Step 1: Find two most recent Excel files ===
# xlsx_files = [
#     os.path.join(input_folder, f)
#     for f in os.listdir(input_folder)
#     if f.lower().endswith(".xlsx")
# ]
# xlsx_files.sort(key=os.path.getmtime, reverse=True)

# if len(xlsx_files) < 2:
#     raise Exception("Need at least two Excel files to compare.")

# new_file, old_file = xlsx_files[0], xlsx_files[1]
# print(f"Comparing:\nNEW: {os.path.basename(new_file)}\nOLD: {os.path.basename(old_file)}")

# # === Step 2: Load both Excel files ===
# new_excel = pd.read_excel(new_file, sheet_name=None)
# old_excel = pd.read_excel(old_file, sheet_name=None)

# # === Step 3: Compare by eventId ===
# differences = []

# for sheet in new_excel:
#     if sheet not in old_excel:
#         continue

#     new_df = new_excel[sheet].copy()
#     old_df = old_excel[sheet].copy()

#     if 'eventId' not in new_df.columns or 'eventId' not in old_df.columns:
#         continue

#     # Normalize eventId as string
#     new_df['eventId'] = new_df['eventId'].astype(str)
#     old_df['eventId'] = old_df['eventId'].astype(str)

#     new_df.set_index('eventId', inplace=True, drop=False)
#     old_df.set_index('eventId', inplace=True, drop=False)

#     # Find added eventIds
#     added_ids = set(new_df.index) - set(old_df.index)
#     if added_ids:
#         added = new_df.loc[list(added_ids)].copy()
#         added['ChangeType'] = 'Added'
#         added['Sheet'] = sheet
#         differences.append(added)

#     # Find modified eventIds
#     common_ids = set(new_df.index).intersection(old_df.index)
#     for eid in common_ids:
#         new_row = new_df.loc[eid]
#         old_row = old_df.loc[eid]
#         if not new_row.equals(old_row):
#             changed = new_row.copy()
#             changed['ChangeType'] = 'Modified'
#             changed['Sheet'] = sheet
#             if isinstance(changed, pd.Series):
#                 differences.append(changed.to_frame().T)
#             else:
#                 differences.append(changed)

# # === Step 4: Save differences to Excel ===
# if differences:
#     all_diffs = pd.concat(differences, ignore_index=True)
#     with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
#         all_diffs.to_excel(writer, sheet_name="Differences", index=False)
#     print(f"\n✅ Differences saved to:\n{output_file}")
# else:
#     print("\n✅ No differences found between the two files.")


# import os
# import pandas as pd
# from datetime import datetime
# from openpyxl import load_workbook
# from openpyxl.styles import PatternFill

# # === Folder paths ===
# input_folder = r"C:\Users\hgada\Downloads\Media_Star_XML\File Output"
# output_folder = r"C:\Users\hgada\Downloads\Media_Star_XML\Differences"
# os.makedirs(output_folder, exist_ok=True)

# # Generate timestamped output file
# timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
# output_file = os.path.join(output_folder, f"Difference_{timestamp}.xlsx")

# # === Step 1: Get last two Excel files ===
# xlsx_files = [
#     os.path.join(input_folder, f)
#     for f in os.listdir(input_folder)
#     if f.lower().endswith(".xlsx")
# ]
# xlsx_files.sort(key=os.path.getmtime, reverse=True)

# if len(xlsx_files) < 2:
#     raise Exception("Need at least two Excel files to compare.")

# new_file, old_file = xlsx_files[0], xlsx_files[1]
# print(f"Comparing:\nNEW: {os.path.basename(new_file)}\nOLD: {os.path.basename(old_file)}")

# # === Step 2: Load Excel files ===
# new_excel = pd.read_excel(new_file, sheet_name=None)
# old_excel = pd.read_excel(old_file, sheet_name=None)

# # === Step 3: Compare eventId-based rows ===
# diff_rows = []

# for sheet in new_excel:
#     if sheet not in old_excel:
#         continue

#     new_df = new_excel[sheet].copy()
#     old_df = old_excel[sheet].copy()

#     if 'eventId' not in new_df.columns or 'eventId' not in old_df.columns:
#         continue

#     new_df['eventId'] = new_df['eventId'].astype(str)
#     old_df['eventId'] = old_df['eventId'].astype(str)

#     new_df.set_index('eventId', inplace=True, drop=False)
#     old_df.set_index('eventId', inplace=True, drop=False)

#     # New rows
#     added_ids = set(new_df.index) - set(old_df.index)
#     if added_ids:
#         added = new_df.loc[list(added_ids)].copy()
#         added['ChangeType'] = 'Added'
#         added['Sheet'] = sheet
#         diff_rows.append(added)

#     # Modified rows
#     common_ids = set(new_df.index).intersection(old_df.index)
#     for eid in common_ids:
#         new_row = new_df.loc[eid]
#         old_row = old_df.loc[eid]
#         if not new_row.equals(old_row):
#             diff = new_row.copy()
#             diff['ChangeType'] = 'Modified'
#             diff['Sheet'] = sheet
#             if isinstance(diff, pd.Series):
#                 diff_rows.append(diff.to_frame().T)
#             else:
#                 diff_rows.append(diff)

# # === Step 4: Write differences ===
# if diff_rows:
#     combined_df = pd.concat(diff_rows, ignore_index=True)
#     combined_df.to_excel(output_file, sheet_name="Differences", index=False)

#     # === Step 5: Re-open to apply highlight styling ===
#     wb = load_workbook(output_file)
#     ws = wb["Differences"]

#     # Reload old and new for comparison
#     new_df_full = pd.concat(new_excel.values(), ignore_index=True)
#     old_df_full = pd.concat(old_excel.values(), ignore_index=True)
#     new_df_full['eventId'] = new_df_full['eventId'].astype(str)
#     old_df_full['eventId'] = old_df_full['eventId'].astype(str)
#     old_df_full.set_index('eventId', inplace=True, drop=False)

#     yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

#     headers = [cell.value for cell in ws[1]]
#     col_map = {header: idx+1 for idx, header in enumerate(headers)}

#     for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
#         eid = str(row[col_map['eventId'] - 1].value)
#         change_type = row[col_map['ChangeType'] - 1].value
#         if change_type != "Modified" or eid not in old_df_full.index:
#             continue

#         old_row = old_df_full.loc[eid]
#         for col_name, col_idx in col_map.items():
#             if col_name in ['eventId', 'ChangeType', 'Sheet']:
#                 continue
#             cell = row[col_idx - 1]
#             old_val = str(old_row.get(col_name, "")).strip()
#             new_val = str(cell.value).strip() if cell.value is not None else ""
#             if old_val != new_val:
#                 cell.fill = yellow_fill

#     wb.save(output_file)
#     print(f"\n✅ Differences with highlights saved to:\n{output_file}")
# else:
#     print("\n✅ No differences found.")


# import os
# import pandas as pd
# from datetime import datetime
# from openpyxl import load_workbook
# from openpyxl.styles import PatternFill

# # === Paths ===
# input_folder = r"C:\Users\hgada\Downloads\Media_Star_XML\File Output"
# output_folder = r"C:\Users\hgada\Downloads\Media_Star_XML\Differences"
# os.makedirs(output_folder, exist_ok=True)

# timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
# output_file = os.path.join(output_folder, f"Difference_{timestamp}.xlsx")

# # === Step 1: Get the two latest Excel files ===
# xlsx_files = [os.path.join(input_folder, f) for f in os.listdir(input_folder) if f.endswith(".xlsx")]
# xlsx_files.sort(key=os.path.getmtime, reverse=True)

# if len(xlsx_files) < 2:
#     raise Exception("Need at least two Excel files to compare.")

# new_file, old_file = xlsx_files[0], xlsx_files[1]
# print(f"Comparing:\nNEW → {os.path.basename(new_file)}\nOLD → {os.path.basename(old_file)}")

# # === Step 2: Load both files ===
# new_excel = pd.read_excel(new_file, sheet_name=None)
# old_excel = pd.read_excel(old_file, sheet_name=None)

# diff_rows = []
# highlight_coords = []

# # === Step 3: Compare tab-by-tab using eventId ===
# for sheet in new_excel:
#     if sheet not in old_excel:
#         continue

#     new_df = new_excel[sheet].copy()
#     old_df = old_excel[sheet].copy()

#     if 'eventId' not in new_df.columns or 'eventId' not in old_df.columns:
#         continue

#     new_df['eventId'] = new_df['eventId'].astype(str)
#     old_df['eventId'] = old_df['eventId'].astype(str)

#     new_df.set_index('eventId', inplace=True, drop=False)
#     old_df.set_index('eventId', inplace=True, drop=False)

#     # 1. Added eventIds
#     added_ids = set(new_df.index) - set(old_df.index)
#     if added_ids:
#         added = new_df.loc[list(added_ids)].copy()
#         added['ChangeType'] = 'Added'
#         added['Sheet'] = sheet
#         diff_rows.append(added)

#     # 2. Modified eventIds
#     common_ids = set(new_df.index).intersection(old_df.index)
#     for eid in common_ids:
#         new_row = new_df.loc[eid]
#         old_row = old_df.loc[eid]

#         if not new_row.equals(old_row):
#             row_data = new_row.copy()
#             row_data['ChangeType'] = 'Modified'
#             row_data['Sheet'] = sheet
#             diff_rows.append(row_data.to_frame().T)

#             # Track cell differences for highlighting
#             for col in new_row.index:
#                 if col in ['eventId', 'Sheet', 'ChangeType']:
#                     continue
#                 if pd.isna(new_row[col]) and pd.isna(old_row.get(col)):
#                     continue
#                 if str(new_row[col]) != str(old_row.get(col)):
#                     highlight_coords.append((eid, col))

# # === Step 4: Save differences to Excel ===
# if diff_rows:
#     combined_df = pd.concat(diff_rows, ignore_index=True)
#     combined_df.to_excel(output_file, sheet_name="Differences", index=False)

#     # === Step 5: Highlight changed cells only ===
#     wb = load_workbook(output_file)
#     ws = wb["Differences"]
#     yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

#     headers = [cell.value for cell in ws[1]]
#     col_index = {col: idx + 1 for idx, col in enumerate(headers)}
#     eventId_idx = col_index['eventId']

#     # Find and highlight specific cells by row+col
#     for row_num in range(2, ws.max_row + 1):
#         eid = str(ws.cell(row=row_num, column=eventId_idx).value)
#         for col_name in headers:
#             if (eid, col_name) in highlight_coords:
#                 col_num = col_index[col_name]
#                 ws.cell(row=row_num, column=col_num).fill = yellow_fill

#     wb.save(output_file)
#     print(f"\n✅ Differences saved and highlighted at:\n{output_file}")
# else:
#     print("\n✅ No differences found. No output file created.")


import os
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# === Folder paths ===
input_folder = r"C:\Users\hgada\Downloads\Media_Star_XML\File Output"
output_folder = r"C:\Users\hgada\Downloads\Media_Star_XML\Differences"
os.makedirs(output_folder, exist_ok=True)

# === Output file setup ===
timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
output_file = os.path.join(output_folder, f"Difference_{timestamp}.xlsx")

# === Step 1: Get the two most recent Excel files ===
xlsx_files = [os.path.join(input_folder, f) for f in os.listdir(input_folder) if f.endswith(".xlsx")]
xlsx_files.sort(key=os.path.getmtime, reverse=True)

if len(xlsx_files) < 2:
    raise Exception("Need at least two Excel files to compare.")

new_file, old_file = xlsx_files[0], xlsx_files[1]
print(f"Comparing:\nNEW → {os.path.basename(new_file)}\nOLD → {os.path.basename(old_file)}")

# === Step 2: Load Excel files ===
new_excel = pd.read_excel(new_file, sheet_name=None)
old_excel = pd.read_excel(old_file, sheet_name=None)

diff_rows = []
highlight_coords = []

# === Step 3: Compare tab-by-tab using eventId ===
for sheet in new_excel:
    if sheet not in old_excel:
        continue

    new_df = new_excel[sheet].copy()
    old_df = old_excel[sheet].copy()

    if 'eventId' not in new_df.columns or 'eventId' not in old_df.columns:
        continue

    new_df['eventId'] = new_df['eventId'].astype(str)
    old_df['eventId'] = old_df['eventId'].astype(str)

    new_df.set_index('eventId', inplace=True, drop=False)
    old_df.set_index('eventId', inplace=True, drop=False)

    # 1. Added eventIds
    added_ids = set(new_df.index) - set(old_df.index)
    if added_ids:
        added = new_df.loc[list(added_ids)].copy()
        added['ChangeType'] = 'Added'
        added['Sheet'] = sheet
        diff_rows.append(added)

    # 2. Modified eventIds (ignore case-only changes)
    common_ids = set(new_df.index).intersection(old_df.index)
    for eid in common_ids:
        new_row = new_df.loc[eid]
        old_row = old_df.loc[eid]

        modified_cols = []
        for col in new_row.index:
            if col in ['eventId', 'Sheet', 'ChangeType']:
                continue

            val_new = str(new_row[col]).strip().lower()
            val_old = str(old_row.get(col, "")).strip().lower()

            if pd.isna(new_row[col]) and pd.isna(old_row.get(col)):
                continue

            if val_new != val_old:
                modified_cols.append(col)

        if modified_cols:
            row_data = new_row.copy()
            row_data['ChangeType'] = 'Modified'
            row_data['Sheet'] = sheet
            diff_rows.append(row_data.to_frame().T)

            for col in modified_cols:
                highlight_coords.append((eid, col))

# === Step 4: Save differences to Excel ===
if diff_rows:
    combined_df = pd.concat(diff_rows, ignore_index=True)
    combined_df.to_excel(output_file, sheet_name="Differences", index=False)

    # === Step 5: Highlight changed cells ===
    wb = load_workbook(output_file)
    ws = wb["Differences"]
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    headers = [cell.value for cell in ws[1]]
    col_index = {col: idx + 1 for idx, col in enumerate(headers)}
    eventId_idx = col_index['eventId']

    for row_num in range(2, ws.max_row + 1):
        eid = str(ws.cell(row=row_num, column=eventId_idx).value)
        for col_name in headers:
            if (eid, col_name) in highlight_coords:
                col_num = col_index[col_name]
                ws.cell(row=row_num, column=col_num).fill = yellow_fill

    wb.save(output_file)
    print(f"\n✅ Differences saved and highlighted at:\n{output_file}")
else:
    print("\n✅ No differences found. No output file created.")
